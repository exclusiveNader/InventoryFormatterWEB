
import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Inventory Formatter Hub", layout="centered")

st.image("logo.png", width=220)
st.title("🛠️ Formatter Hub")

with st.expander("📘 How to Use"):
    st.markdown("""
    1. Choose a tab for the type of report you want to format  
    2. Go to Leaflink and download the report file of your choosing
    3. 📁 Upload your raw report file (.csv or .xlsx)
    4. 📥 Click the green button to download your formatted Excel file  

    **Works on Mac, PC, and mobile browsers.**
    """)

tabs = st.tabs(["📦 Inventory Formatter", "📈 Products Sold Formatter", "📝 Order Report Formatter"])

# ───── Inventory Formatter ─────
with tabs[0]:
    st.subheader("📦 Inventory Formatter")
    uploaded_file = st.file_uploader("Upload inventory file", type=["csv", "xlsx"], key="inv")

    def clean_price(value):
        if pd.isna(value): return 0.0
        if isinstance(value, (int, float)): return float(value)
        cleaned = re.sub(r"[^0-9.]", "", str(value))
        try: return float(cleaned)
        except: return 0.0

    def process_inventory(df_raw):
        column_map = {
            "Wholesale Price ($)": "Wholesale Price",
            "Wholesale Price": "Wholesale Price",
            "Available Inventory (Units)": "Available Inventory (Units)"
        }
        required = ["Name", "Wholesale Price", "Brand", "Product Line", "Classification", "Listing State", "Available Inventory (Units)"]
        for alt, canonical in column_map.items():
            if alt in df_raw.columns and canonical not in df_raw.columns:
                df_raw.rename(columns={alt: canonical}, inplace=True)
        for col in required:
            if col not in df_raw.columns:
                raise ValueError(f"Missing column: {col}")

        df = df_raw[required].copy()
        df["Wholesale Price"] = df["Wholesale Price"].apply(clean_price).round(2)
        df["Product Line"] = df["Product Line"].fillna("Uncategorized")
        df.sort_values(["Brand", "Product Line", "Name"], inplace=True, ignore_index=True)

        grouped = []
        for (brand, line), group in df.groupby(["Brand", "Product Line"]):
            grouped.append(group)
            total_units = group["Available Inventory (Units)"].sum()
            total_row = {
                "Name": f"TOTAL - {line}",
                "Wholesale Price": "",
                "Brand": brand,
                "Product Line": line,
                "Classification": "",
                "Listing State": "",
                "Available Inventory (Units)": total_units
            }
            grouped.append(pd.DataFrame([total_row]))
            grouped.append(pd.DataFrame([{}]))  # blank line

        return pd.concat(grouped, ignore_index=True)

    def to_excel(df):
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Formatted", index=False)
            sheet = writer.sheets["Formatted"]

            red_bold = Font(color="FF0000", bold=True)
            bold = Font(bold=True)
            center = Alignment(horizontal="center", vertical="center")

            for col_num, col_name in enumerate(df.columns, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.font = red_bold
                cell.alignment = center
                col_letter = get_column_letter(col_num)
                sheet.column_dimensions[col_letter].width = 30 if col_name != "Name" else 60

            for row in range(2, sheet.max_row + 1):
                name_val = sheet.cell(row=row, column=1).value
                if isinstance(name_val, str) and name_val.startswith("TOTAL"):
                    for col in range(1, len(df.columns) + 1):
                        sheet.cell(row=row, column=col).font = bold

            sheet.auto_filter.ref = sheet.dimensions
            sheet.freeze_panes = "A2"

        output.seek(0)
        return output

    if uploaded_file:
        try:
            if uploaded_file.name.endswith(".csv"):
                df_raw = pd.read_csv(uploaded_file)
            else:
                df_raw = pd.read_excel(uploaded_file)

            df_formatted = process_inventory(df_raw)
            xlsx_data = to_excel(df_formatted)

            st.success("✅ File formatted successfully!")
            st.download_button("📥 Download Inventory Excel", xlsx_data, file_name="Formatted_Inventory.xlsx")

        except Exception as e:
            st.error(f"⚠️ Error: {e}")

# ───── Products Sold Formatter ─────
with tabs[1]:
    st.subheader("📈 Products Sold Formatter")
    uploaded_file = st.file_uploader("Upload products sold file", type=["csv", "xlsx"], key="sold")

    if uploaded_file:
        try:
            if uploaded_file.name.endswith(".csv"):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)

            df.columns = [c.strip() for c in df.columns]
            df["Wholesale Price"] = df["Wholesale Price"].replace('[\$,]', '', regex=True).astype(float)
            df["Product Line"] = df["Product Line"].fillna("Uncategorized")

            df_out = df[[
                "Product", "Brand", "Product Line",
                "Shelf Inventory", "Wholesale Price",
                "Amount Sold (Units)", "Amount Sold (Cases)"
            ]].copy()

            df_out.sort_values(["Brand", "Product Line", "Product"], inplace=True, ignore_index=True)

            grouped_output = []
            for (brand, line), group in df_out.groupby(["Brand", "Product Line"]):
                grouped_output.append(group)

                total_units = group["Amount Sold (Units)"].sum()
                total_dollars = (group["Wholesale Price"] * group["Amount Sold (Units)"]).sum()

                total_unit_row = {
                    "Amount Sold (Units)": f"Total Units Sold: {int(total_units)}"
                }
                total_dollars_row = {
                    "Amount Sold (Units)": f"Total Dollars Sold: ${total_dollars:,.2f}"
                }

                grouped_output.append(pd.DataFrame([total_unit_row]))
                grouped_output.append(pd.DataFrame([{}]))

            df_final = pd.concat(grouped_output, ignore_index=True)

            def to_excel_sold(df):
                output = BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    df.to_excel(writer, sheet_name="Formatted", index=False)
                    sheet = writer.sheets["Formatted"]

                    red_bold = Font(color="FF0000", bold=True)
                    bold = Font(bold=True)
                    center = Alignment(horizontal="center", vertical="center")

                    for col_num, col_name in enumerate(df.columns, 1):
                        cell = sheet.cell(row=1, column=col_num)
                        cell.font = red_bold
                        cell.alignment = center
                        col_letter = get_column_letter(col_num)
                        sheet.column_dimensions[col_letter].width = 28 if col_name != "Product" else 60

                    sheet.auto_filter.ref = sheet.dimensions
                    sheet.freeze_panes = "A2"

                output.seek(0)
                return output

            xlsx_sold = to_excel_sold(df_final)
            st.success("✅ Products Sold formatted!")
            st.download_button("📥 Download Products Sold Excel", data=xlsx_sold, file_name="Formatted_Products_Sold.xlsx")

        except Exception as e:
            st.error(f"⚠️ Error: {e}")

# ───── Order Report Formatter ─────
with tabs[2]:
    st.subheader("📝 Order Report Formatter")
    uploaded_file = st.file_uploader("Upload order report file", type=["csv", "xlsx"], key="order")

    if uploaded_file:
        try:
            if uploaded_file.name.endswith(".csv"):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)

            df.columns = [c.strip() for c in df.columns]

            # Fix Unit Price
            if "Unit Price" in df.columns:
                df["Unit Price"] = df["Unit Price"].replace('[\\$,]', '', regex=True).astype(float)
            else:
                df["Unit Price"] = 0.0

            # Fix Quantity
            if "Quantity" in df.columns:
                df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0).astype(int)
            else:
                df["Quantity"] = 0

            # Fix Total Price or calculate it
            if "Total Price" in df.columns:
                df["Total Price"] = df["Total Price"].replace('[\\$,]', '', regex=True).astype(float)
            else:
                df["Total Price"] = df["Unit Price"] * df["Quantity"]

            df_out = df[["Buyer Name", "Brand", "Quantity", "Total Price"]].copy()
            df_out.columns = ["Customer", "Brand", "Qty (Units)", "Line Item Total"]
            df_out.sort_values(["Customer", "Brand"], inplace=True)

            # Group and subtotal
            grouped_output = []
            for customer, group in df_out.groupby("Customer"):
                grouped_output.append(group)

                total_qty = group["Qty (Units)"].sum()
                total_price = group["Line Item Total"].sum()

                total_row = {
                    "Customer": f"TOTAL - {customer}",
                    "Qty (Units)": total_qty,
                    "Line Item Total": f"${total_price:,.2f}"
                }
                grouped_output.append(pd.DataFrame([total_row]))
                grouped_output.append(pd.DataFrame([{}]))

            df_final = pd.concat(grouped_output, ignore_index=True)

            def to_excel_order(df):
                output = BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    df.to_excel(writer, sheet_name="Formatted", index=False)
                    sheet = writer.sheets["Formatted"]

                    red_bold = Font(color="FF0000", bold=True)
                    bold = Font(bold=True)
                    center = Alignment(horizontal="center", vertical="center")

                    for col_num, col_name in enumerate(df.columns, 1):
                        cell = sheet.cell(row=1, column=col_num)
                        cell.font = red_bold
                        cell.alignment = center
                        col_letter = get_column_letter(col_num)
                        sheet.column_dimensions[col_letter].width = 28 if col_name != "Customer" else 60

                    for row in range(2, sheet.max_row + 1):
                        val = sheet.cell(row=row, column=1).value
                        if isinstance(val, str) and val.startswith("TOTAL -"):
                            for col in range(1, len(df.columns) + 1):
                                sheet.cell(row=row, column=col).font = bold

                    sheet.auto_filter.ref = sheet.dimensions
                    sheet.freeze_panes = "A2"

                output.seek(0)
                return output

            xlsx_order = to_excel_order(df_final)
            st.success("✅ Order Report formatted!")
            st.download_button("📥 Download Order Report Excel", data=xlsx_order, file_name="Formatted_Order_Report.xlsx")

        except Exception as e:
            st.error(f"⚠️ Error: {e}")

